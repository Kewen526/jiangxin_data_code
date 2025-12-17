#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
江鑫数据报表生成系统
支持生成：日报、周报、月报、自定义报表
"""

import mysql.connector
from mysql.connector import pooling
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
import warnings
warnings.filterwarnings('ignore')

# ==================== 数据库连接池配置 ====================
DB_CONFIG = {
    'host': '8.146.210.145',
    'port': 3306,
    'user': 'root',
    'password': 'Kewen888@',
    'database': 'jx_data_info',  # 数据库名称
    'charset': 'utf8mb4',
    'use_unicode': True,
    'autocommit': True
}

# 创建全局连接池（单例模式）
CONNECTION_POOL = mysql.connector.pooling.MySQLConnectionPool(
    pool_name="jx_pool",
    pool_size=20,  # 连接池大小
    pool_reset_session=True,
    **DB_CONFIG
)


# ==================== 辅助函数 ====================
def get_shop_info_mapping(accounts=None):
    """
    获取门店信息映射
    参数:
        accounts: list, 可选，账号列表（platform_accounts.account的值），如果提供则只查询这些账号
    返回: dict {shop_id: {'operator': '', 'sales': '', 'city': ''}}
    """
    conn = CONNECTION_POOL.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 1. 查询 platform_accounts 和 saas_users
        sql = """
        SELECT
            pa.account,
            pa.stores_json,
            pa.sales_name,
            pa.city_name,
            pa.operator_id,
            su.name as operator_name
        FROM platform_accounts pa
        LEFT JOIN saas_users su ON pa.operator_id = su.manager_id
        WHERE pa.stores_json IS NOT NULL
        """

        # 如果指定了accounts列表，添加过滤条件
        params = []
        if accounts:
            placeholders = ','.join(['%s'] * len(accounts))
            sql += f" AND pa.account IN ({placeholders})"
            params = accounts

        cursor.execute(sql, params)
        account_results = cursor.fetchall()

        # 2. 解析 stores_json 构建映射
        shop_mapping = {}

        for account in account_results:
            stores_json = account.get('stores_json')
            sales_name = account.get('sales_name', '')
            city_name = account.get('city_name', '')
            operator_name = account.get('operator_name', '')

            if stores_json:
                try:
                    # 解析 JSON（可能是字符串或已经是对象）
                    if isinstance(stores_json, str):
                        stores = json.loads(stores_json)
                    else:
                        stores = stores_json

                    # 遍历门店列表
                    if isinstance(stores, list):
                        for store in stores:
                            if isinstance(store, dict):
                                shop_id = str(store.get('shop_id', ''))
                                if shop_id:
                                    shop_mapping[shop_id] = {
                                        'operator': operator_name or '',
                                        'sales': sales_name or '',
                                        'city': city_name or ''
                                    }
                except (json.JSONDecodeError, TypeError):
                    # 忽略解析错误
                    pass

        return shop_mapping

    finally:
        cursor.close()
        conn.close()


def get_region_info_mapping(accounts=None):
    """
    获取商圈信息映射
    参数:
        accounts: list, 可选，账号列表（platform_accounts.account的值），如果提供则只查询这些账号
    返回: dict {shop_id: {'city': '', 'district': '', 'region': ''}}
    数据来源: platform_accounts.compareRegions_json
    """
    conn = CONNECTION_POOL.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        sql = """
        SELECT
            pa.stores_json,
            pa.compareRegions_json
        FROM platform_accounts pa
        WHERE pa.stores_json IS NOT NULL
        """

        # 如果指定了accounts列表，添加过滤条件
        params = []
        if accounts:
            placeholders = ','.join(['%s'] * len(accounts))
            sql += f" AND pa.account IN ({placeholders})"
            params = accounts

        cursor.execute(sql, params)
        account_results = cursor.fetchall()

        region_mapping = {}

        for account in account_results:
            stores_json = account.get('stores_json')
            regions_json = account.get('compareRegions_json')

            if stores_json and regions_json:
                try:
                    # 解析门店列表
                    if isinstance(stores_json, str):
                        stores = json.loads(stores_json)
                    else:
                        stores = stores_json

                    # 解析商圈数据
                    if isinstance(regions_json, str):
                        regions = json.loads(regions_json)
                    else:
                        regions = regions_json

                    # 遍历门店，匹配商圈信息
                    if isinstance(stores, list):
                        for store in stores:
                            if isinstance(store, dict):
                                shop_id = str(store.get('shop_id', ''))
                                if shop_id and isinstance(regions, dict):
                                    # regions 格式可能是 {shop_id: {city, district, region}}
                                    # 或者是列表格式
                                    if shop_id in regions:
                                        region_data = regions[shop_id]
                                        region_mapping[shop_id] = {
                                            'city': region_data.get('city', ''),
                                            'district': region_data.get('district', ''),
                                            'region': region_data.get('region', '')
                                        }
                                    elif isinstance(regions, list):
                                        for r in regions:
                                            if str(r.get('shop_id', '')) == shop_id:
                                                region_mapping[shop_id] = {
                                                    'city': r.get('city', ''),
                                                    'district': r.get('district', ''),
                                                    'region': r.get('region', '')
                                                }
                                                break
                except (json.JSONDecodeError, TypeError):
                    pass

        return region_mapping

    finally:
        cursor.close()
        conn.close()


def get_coupon_orders_last_7days(shop_id, report_date):
    """
    获取近7天优惠码订单总数
    参数:
        shop_id: 门店ID
        report_date: 报表日期 (str 'YYYY-MM-DD')
    返回: int 近7天优惠码订单总数
    数据来源: kewen_daily_report.coupon_pay_order_count
    """
    conn = CONNECTION_POOL.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 计算7天前的日期
        end_date = datetime.strptime(report_date, '%Y-%m-%d')
        start_date = end_date - timedelta(days=6)

        sql = """
        SELECT COALESCE(SUM(coupon_pay_order_count), 0) as total
        FROM kewen_daily_report
        WHERE shop_id = %s
          AND report_date BETWEEN %s AND %s
        """

        cursor.execute(sql, (shop_id, start_date.strftime('%Y-%m-%d'), report_date))
        result = cursor.fetchone()

        return int(result['total']) if result and result['total'] else 0

    finally:
        cursor.close()
        conn.close()


def get_ad_orders_last_7days(shop_id, report_date):
    """
    获取近7天广告单总数
    参数:
        shop_id: 门店ID
        report_date: 报表日期 (str 'YYYY-MM-DD')
    返回: int 近7天广告单总数
    数据来源: store_stats.ad_order_count
    """
    conn = CONNECTION_POOL.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 计算7天前的日期
        end_date = datetime.strptime(report_date, '%Y-%m-%d')
        start_date = end_date - timedelta(days=6)

        sql = """
        SELECT COALESCE(SUM(ad_order_count), 0) as total
        FROM store_stats
        WHERE store_id = %s
          AND date BETWEEN %s AND %s
        """

        cursor.execute(sql, (shop_id, start_date.strftime('%Y-%m-%d'), report_date))
        result = cursor.fetchone()

        return int(result['total']) if result and result['total'] else 0

    finally:
        cursor.close()
        conn.close()


def clean_sheet_name(name, max_length=31):
    """
    清理 Sheet 名称，符合 Excel 规范
    - 最大 31 字符
    - 不能包含: \ / * ? : [ ]
    """
    if not name:
        return "Sheet"

    # 替换非法字符
    illegal_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in illegal_chars:
        name = name.replace(char, '')

    # 截断到最大长度
    if len(name) > max_length:
        name = name[:max_length]

    return name or "Sheet"


def apply_border(ws, min_row, max_row, min_col, max_col):
    """应用边框样式"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border


# ==================== 核心功能：生成日报 ====================
def generate_daily_report(report_date, accounts=None, output_filename=None):
    """
    生成日报
    - Sheet 1: "汇总" - 横向表格，每行一个门店
    - Sheet 2-N: 门店详细 - 竖向表格

    参数:
        report_date: str, 报表日期，格式: 'YYYY-MM-DD'
        accounts: list, 可选，门店账号列表（platform_accounts.account的值），如["13718175572a","19318574226a"]，如果提供则只生成这些账号的日报
        output_filename: str, 输出文件名，默认自动生成

    返回:
        str: 生成的文件路径
    """
    # 1. 如果指定了accounts，先获取对应的shop_id列表
    shop_ids_filter = None
    if accounts:
        print(f"正在查询指定账号的门店信息: {accounts}")
        conn_temp = CONNECTION_POOL.get_connection()
        cursor_temp = conn_temp.cursor(dictionary=True)
        try:
            placeholders = ','.join(['%s'] * len(accounts))
            sql_accounts = f"""
            SELECT stores_json, compareRegions_json
            FROM platform_accounts
            WHERE account IN ({placeholders})
            """
            cursor_temp.execute(sql_accounts, accounts)
            account_data = cursor_temp.fetchall()

            # 从stores_json中提取shop_id
            shop_ids_filter = []
            for acc in account_data:
                stores_json = acc.get('stores_json')
                if stores_json:
                    try:
                        if isinstance(stores_json, str):
                            stores = json.loads(stores_json)
                        else:
                            stores = stores_json

                        if isinstance(stores, list):
                            for store in stores:
                                if isinstance(store, dict):
                                    shop_id = str(store.get('shop_id', ''))
                                    if shop_id:
                                        shop_ids_filter.append(shop_id)
                    except (json.JSONDecodeError, TypeError):
                        pass

            print(f"找到 {len(shop_ids_filter)} 个门店")
        finally:
            cursor_temp.close()
            conn_temp.close()

    # 2. 获取门店信息映射
    print("正在加载门店信息...")
    shop_mapping = get_shop_info_mapping(accounts)
    region_mapping = get_region_info_mapping(accounts)

    # 3. 从连接池获取连接
    conn = CONNECTION_POOL.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # SQL 查询：关联 kewen_daily_report + promotion_daily_report + store_stats
        sql = """
        SELECT
            k.report_date,
            k.shop_id,
            k.shop_name,
            k.exposure_users,
            k.visit_users,
            k.order_users,
            k.verify_person_count as verify_users,
            k.order_coupon_count,
            k.verify_coupon_count,
            k.promotion_cost,
            k.new_good_review_count,
            k.new_review_count,
            k.new_collect_users,
            k.consult_users,
            k.intent_rate,
            k.order_sale_amount,
            k.verify_sale_amount,
            k.verify_after_discount,
            p.view_phone_count as phone_clicks,
            p.view_address_count as address_clicks,
            p.click_avg_price,
            p.order_count as promotion_order_count,
            s.order_user_rank,
            s.verify_amount_rank,
            s.checkin_count,
            s.ad_balance,
            s.ad_order_count,
            s.is_force_offline
        FROM kewen_daily_report k
        LEFT JOIN promotion_daily_report p
            ON k.shop_id = p.shop_id AND k.report_date = p.report_date
        LEFT JOIN store_stats s
            ON k.shop_id = s.store_id AND k.report_date = s.date
        WHERE k.report_date = %s
        """

        # 添加shop_id过滤条件
        params = [report_date]
        if shop_ids_filter:
            placeholders = ','.join(['%s'] * len(shop_ids_filter))
            sql += f" AND k.shop_id IN ({placeholders})"
            params.extend(shop_ids_filter)

        sql += " ORDER BY k.shop_id"

        cursor.execute(sql, params)
        rows = cursor.fetchall()

        if not rows:
            print(f"警告：{report_date} 没有数据")
            return None

        # 3. 创建 Excel 工作簿
        wb = openpyxl.Workbook()

        # ==================== Sheet 1: 汇总 ====================
        ws_summary = wb.active
        ws_summary.title = "汇总"

        # 格式化日期
        date_obj = datetime.strptime(report_date, '%Y-%m-%d')
        weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
        weekday = weekday_names[date_obj.weekday()]
        date_str = date_obj.strftime('%m月%d日')
        date_short = date_obj.strftime('%m/%d')

        # 汇总表头
        summary_headers = [
            '星期', '日期', '序号', '运营', '城市', '销售', '门店',
            '曝光人数', '访问人数', '下单人数', '核销人数', '下单券数', '核销券数',
            '电话点击', '地址点击', '推广通消耗', '好评', '意向转化率',
            '下单售价金额', '核销售价金额', '优惠后核销金额',
            '下单人数商圈排名', '核销金额商圈排名'
        ]
        ws_summary.append(summary_headers)

        # 汇总表头样式
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, size=10)
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 用于处理重名 Sheet
        sheet_names_used = {}

        # 4. 为每个门店写入汇总行 + 创建详细Sheet
        for idx, row in enumerate(rows, start=1):
            shop_id = str(row['shop_id'])
            shop_name = row['shop_name'] or f'门店{shop_id}'

            # 从映射中获取运营、城市、销售
            shop_info = shop_mapping.get(shop_id, {})
            operator = shop_info.get('operator', '')
            sales = shop_info.get('sales', '')
            city = shop_info.get('city', '')

            # 获取商圈信息
            region_info = region_mapping.get(shop_id, {})
            region_city = region_info.get('city', city)
            region_district = region_info.get('district', '')
            region_name = region_info.get('region', '')

            # 格式化商圈排名
            order_rank = row['order_user_rank']
            verify_rank = row['verify_amount_rank']
            order_rank_str = f"第{order_rank}名" if order_rank and order_rank < 100 else ("大于100名" if order_rank and order_rank >= 100 else "--")
            verify_rank_str = f"第{verify_rank}名" if verify_rank and verify_rank < 100 else ("大于100名" if verify_rank and verify_rank >= 100 else "--")

            # 写入汇总数据行
            summary_row = [
                weekday,
                date_str,
                idx,
                operator,
                city,
                sales,
                shop_name,
                row['exposure_users'] or 0,
                row['visit_users'] or 0,
                row['order_users'] or 0,
                row['verify_users'] or 0,
                row['order_coupon_count'] or 0,
                row['verify_coupon_count'] or 0,
                row['phone_clicks'] or 0,
                row['address_clicks'] or 0,
                round(row['promotion_cost'], 2) if row['promotion_cost'] else 0,
                row['new_good_review_count'] or 0,
                row['intent_rate'] or '0%',
                round(row['order_sale_amount'], 2) if row['order_sale_amount'] else 0,
                round(row['verify_sale_amount'], 2) if row['verify_sale_amount'] else 0,
                round(row['verify_after_discount'], 2) if row['verify_after_discount'] else 0,
                order_rank_str,
                verify_rank_str
            ]
            ws_summary.append(summary_row)

            # ==================== Sheet 2-N: 门店详细（竖向表格）====================
            # 清理 Sheet 名称
            sheet_name = clean_sheet_name(shop_name)

            # 处理重名（添加序号）
            if sheet_name in sheet_names_used:
                sheet_names_used[sheet_name] += 1
                sheet_name = f"{sheet_name[:28]}_{sheet_names_used[sheet_name]}"
            else:
                sheet_names_used[sheet_name] = 1

            # 创建详细 Sheet
            ws_detail = wb.create_sheet(title=sheet_name)

            # 计算达标状态
            order_users = row['order_users'] or 0
            verify_users = row['verify_users'] or 0
            new_review_count = row['new_review_count'] or 0
            new_collect_users = row['new_collect_users'] or 0

            # 留评率 = 新增评价 / 核销人数
            review_rate = (new_review_count / verify_users * 100) if verify_users > 0 else 0
            review_rate_str = f"{review_rate:.1f}%"
            review_qualified = "达标" if review_rate >= 30 else "未达标"

            # 收藏率 = 新增收藏 / 下单人数
            collect_rate = (new_collect_users / order_users * 100) if order_users > 0 else 0
            collect_rate_str = f"{collect_rate:.1f}%"
            collect_qualified = "达标" if collect_rate >= 40 else "未达标"

            # 近7天优惠码订单
            coupon_7days = get_coupon_orders_last_7days(shop_id, report_date)
            coupon_qualified = "达标" if coupon_7days >= 10 else "未达标"

            # 近7天广告单
            ad_7days = get_ad_orders_last_7days(shop_id, report_date)
            ad_qualified = "达标" if ad_7days >= 10 else "未达标"

            # 强制下线状态信息
            is_force_offline = row['is_force_offline'] or 0
            if is_force_offline > 0:
                status_info = f"⚠️ 警告：有{is_force_offline}个团单被强制下线！"
            else:
                status_info = "今天邮件已查看，无违规无异常。"

            # 商圈排名显示
            region_display = f"{region_city}|{region_district}|{region_name}" if region_district else city
            order_rank_display = f"{region_display}:第{order_rank}名" if order_rank and order_rank < 100 else f"{region_display}:大于100名"
            verify_rank_display = f"{region_display}:第{verify_rank}名" if verify_rank and verify_rank < 100 else f"{region_display}:大于100名"

            # 构建竖向表格数据
            detail_data = [
                [shop_name, status_info, ''],
                [f"数据报表", f"日期({date_short})", ''],
                ['【美团点评广告结果数据】', '', ''],
                ['曝光人数：', row['exposure_users'] or 0, ''],
                ['访问人数：', row['visit_users'] or 0, ''],
                ['下单人数：', row['order_users'] or 0, ''],
                ['下单券数：', row['order_coupon_count'] or 0, ''],
                ['核销人数：', row['verify_users'] or 0, ''],
                ['核销券数：', row['verify_coupon_count'] or 0, ''],
                ['电话点击：', row['phone_clicks'] or 0, ''],
                ['地址点击：', row['address_clicks'] or 0, ''],
                ['在线咨询：', row['consult_users'] or 0, ''],
                ['', '', ''],
                ['【店内干预数据】', '', ''],
                ['新增收藏：', row['new_collect_users'] or 0, ''],
                ['新增打卡：', row['checkin_count'] or 0, ''],
                ['新增评价：', row['new_review_count'] or 0, ''],
                ['', '', ''],
                ['【推广通数据】', '', ''],
                ['推广通消耗：', round(row['promotion_cost'], 2) if row['promotion_cost'] else 0, ''],
                ['推广通点击单价：', round(row['click_avg_price'], 2) if row['click_avg_price'] else 0, ''],
                ['推广通下单量：', row['promotion_order_count'] or 0, ''],
                ['推广通余额：', round(row['ad_balance'], 2) if row['ad_balance'] else 0, ''],
                ['', '', ''],
                [f'留评率（30%达标）：', review_rate_str, review_qualified],
                [f'收藏率（40%达标）：', collect_rate_str, collect_qualified],
                [f'近7天优惠码订单是否达标：', coupon_7days, coupon_qualified],
                [f'广告单：', f"7天共{ad_7days}单", ad_qualified],
                ['', '', ''],
                ['下单售价金额：', round(row['order_sale_amount'], 2) if row['order_sale_amount'] else 0, ''],
                ['核销售价金额：', round(row['verify_sale_amount'], 2) if row['verify_sale_amount'] else 0, ''],
                ['下单人数商圈排名：', order_rank_display, ''],
                ['核销金额商圈排名：', verify_rank_display, ''],
                ['', '', ''],
                ['团单被强制下线数量：', is_force_offline, ''],
                ['', '', ''],
                ['运营：', operator, ''],
                ['销售：', sales, ''],
                ['城市：', city, ''],
            ]

            # 写入详细数据
            for row_data in detail_data:
                ws_detail.append(row_data)

            # 设置详细Sheet样式
            ws_detail.column_dimensions['A'].width = 25
            ws_detail.column_dimensions['B'].width = 30
            ws_detail.column_dimensions['C'].width = 15

            # 标题行样式
            ws_detail['A1'].font = Font(bold=True, size=12)
            ws_detail['B1'].font = Font(bold=True, size=10, color="FF0000" if is_force_offline > 0 else "008000")

            # 分类标题样式
            section_rows = [3, 14, 19]  # 【美团点评广告结果数据】等行
            for r in section_rows:
                ws_detail.cell(row=r, column=1).font = Font(bold=True, size=10, color="0066CC")

            # 达标/未达标样式
            qualified_rows = [25, 26, 27, 28]
            for r in qualified_rows:
                cell = ws_detail.cell(row=r, column=3)
                if cell.value == "未达标":
                    cell.font = Font(bold=True, color="FF0000")
                elif cell.value == "达标":
                    cell.font = Font(bold=True, color="008000")

            # 应用边框
            apply_border(ws_detail, 1, len(detail_data), 1, 3)

        # 设置汇总表列宽
        summary_widths = [6, 8, 5, 12, 8, 8, 35, 10, 10, 10, 10, 10, 10, 10, 10, 12, 8, 12, 12, 12, 12, 14, 14]
        for col_idx, width in enumerate(summary_widths, start=1):
            ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

        # 应用汇总表边框
        apply_border(ws_summary, 1, len(rows) + 1, 1, len(summary_headers))

        # 5. 保存文件
        if not output_filename:
            output_filename = f"日报 非餐 {report_date.replace('-', '')} {datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        wb.save(output_filename)
        print(f"✅ 日报生成成功: {output_filename}（共 {len(rows)} 个门店）")
        return output_filename

    finally:
        cursor.close()
        conn.close()


# ==================== 核心功能：生成周报 ====================
def generate_weekly_report(week1_start, week1_end, week2_start, week2_end, output_filename=None):
    """
    生成周报（两周对比）
    - Sheet 1: "汇总" - 每门店8行的横向结构
    - Sheet 2-N: 门店详细 - 竖向31行结构

    参数:
        week1_start: str, 第一周开始日期 'YYYY-MM-DD'
        week1_end: str, 第一周结束日期 'YYYY-MM-DD'
        week2_start: str, 第二周开始日期 'YYYY-MM-DD'
        week2_end: str, 第二周结束日期 'YYYY-MM-DD'
        output_filename: str, 输出文件名
    """
    # 获取门店信息映射
    shop_mapping = get_shop_info_mapping()

    conn = CONNECTION_POOL.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 查询周数据（包含 store_stats）
        sql_week = """
        SELECT
            k.shop_id,
            k.shop_name,
            SUM(k.verify_after_discount) as verify_after_discount,
            SUM(k.exposure_users) as exposure_users,
            SUM(k.visit_users) as visit_users,
            SUM(k.order_users) as order_users,
            SUM(k.order_coupon_count) as order_coupon_count,
            SUM(k.verify_person_count) as verify_users,
            SUM(k.verify_coupon_count) as verify_coupon_count,
            SUM(k.order_sale_amount) as order_sale_amount,
            SUM(k.verify_sale_amount) as verify_sale_amount,
            SUM(k.coupon_pay_order_count) as coupon_orders,
            SUM(p.view_phone_count) as phone_clicks,
            SUM(k.promotion_cost) as promotion_cost,
            SUM(k.promotion_exposure_count) as promotion_exposure,
            SUM(k.promotion_click_count) as promotion_clicks,
            SUM(p.order_count) as promotion_orders,
            SUM(p.view_groupbuy_count) as view_groupbuy,
            SUM(p.view_phone_count) as view_phone,
            SUM(k.consult_users) as consult_users,
            SUM(p.view_address_count) as address_clicks,
            SUM(k.new_collect_users) as new_collect,
            SUM(k.new_good_review_count) as new_good_reviews,
            SUM(k.new_review_count) as new_reviews,
            SUM(s.checkin_count) as checkin_count
        FROM kewen_daily_report k
        LEFT JOIN promotion_daily_report p
            ON k.shop_id = p.shop_id AND k.report_date = p.report_date
        LEFT JOIN store_stats s
            ON k.shop_id = s.store_id AND k.report_date = s.date
        WHERE k.report_date BETWEEN %s AND %s
        GROUP BY k.shop_id, k.shop_name
        ORDER BY k.shop_id
        """

        # 获取第一周数据
        cursor.execute(sql_week, (week1_start, week1_end))
        week1_data = {row['shop_id']: row for row in cursor.fetchall()}

        # 获取第二周数据
        cursor.execute(sql_week, (week2_start, week2_end))
        week2_data = {row['shop_id']: row for row in cursor.fetchall()}

        # 获取所有门店列表
        all_shop_ids = set(week1_data.keys()) | set(week2_data.keys())

        if not all_shop_ids:
            print("警告：没有找到数据")
            return None

        # 创建 Excel
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "汇总"

        # 格式化日期周期
        week1_period = f"{datetime.strptime(week1_start, '%Y-%m-%d').strftime('%Y.%m.%d')}-{datetime.strptime(week1_end, '%Y-%m-%d').strftime('%Y.%m.%d')}"
        week2_period = f"{datetime.strptime(week2_start, '%Y-%m-%d').strftime('%Y.%m.%d')}-{datetime.strptime(week2_end, '%Y-%m-%d').strftime('%Y.%m.%d')}"

        # 辅助函数
        def get_val(data, key, default=0):
            return data.get(key, default) if data else default

        def calc_rate(numerator, denominator):
            if denominator and denominator > 0:
                return round(numerator / denominator * 100, 1)
            return 0

        def calc_avg_price(total, count):
            if count and count > 0:
                return round(total / count, 2)
            return 0

        # 用于处理重名 Sheet
        sheet_names_used = {}

        # 为每个门店生成数据
        for shop_id in sorted(all_shop_ids):
            w1 = week1_data.get(shop_id, {})
            w2 = week2_data.get(shop_id, {})
            shop_name = w2.get('shop_name') or w1.get('shop_name', '未知门店')

            # 获取门店信息
            shop_id_str = str(shop_id)
            shop_info = shop_mapping.get(shop_id_str, {})

            # ==================== 汇总Sheet: 8行/门店 ====================
            # 第一周指标
            w1_verify_discount = get_val(w1, 'verify_after_discount')
            w1_exposure = get_val(w1, 'exposure_users')
            w1_visit = get_val(w1, 'visit_users')
            w1_order_users = get_val(w1, 'order_users')
            w1_order_coupons = get_val(w1, 'order_coupon_count')
            w1_verify_users = get_val(w1, 'verify_users')
            w1_verify_coupons = get_val(w1, 'verify_coupon_count')
            w1_order_amount = get_val(w1, 'order_sale_amount')
            w1_verify_amount = get_val(w1, 'verify_sale_amount')
            w1_coupon_orders = get_val(w1, 'coupon_orders')
            w1_phone_clicks = get_val(w1, 'phone_clicks')

            w1_exposure_rate = f"{calc_rate(w1_visit, w1_exposure)}%"
            w1_order_rate = f"{calc_rate(w1_order_users, w1_visit)}%"
            w1_avg_price = calc_avg_price(w1_verify_discount, w1_verify_users)

            # 第二周指标
            w2_verify_discount = get_val(w2, 'verify_after_discount')
            w2_exposure = get_val(w2, 'exposure_users')
            w2_visit = get_val(w2, 'visit_users')
            w2_order_users = get_val(w2, 'order_users')
            w2_order_coupons = get_val(w2, 'order_coupon_count')
            w2_verify_users = get_val(w2, 'verify_users')
            w2_verify_coupons = get_val(w2, 'verify_coupon_count')
            w2_order_amount = get_val(w2, 'order_sale_amount')
            w2_verify_amount = get_val(w2, 'verify_sale_amount')
            w2_coupon_orders = get_val(w2, 'coupon_orders')
            w2_phone_clicks = get_val(w2, 'phone_clicks')

            w2_exposure_rate = f"{calc_rate(w2_visit, w2_exposure)}%"
            w2_order_rate = f"{calc_rate(w2_order_users, w2_visit)}%"
            w2_avg_price = calc_avg_price(w2_verify_discount, w2_verify_users)

            # 计算差值
            diff_verify_discount = round(w2_verify_discount - w1_verify_discount, 2)
            diff_exposure = w2_exposure - w1_exposure
            diff_visit = w2_visit - w1_visit
            diff_order_users = w2_order_users - w1_order_users
            diff_order_coupons = w2_order_coupons - w1_order_coupons
            diff_verify_users = w2_verify_users - w1_verify_users
            diff_verify_coupons = w2_verify_coupons - w1_verify_coupons
            diff_order_amount = round(w2_order_amount - w1_order_amount, 2)
            diff_verify_amount = round(w2_verify_amount - w1_verify_amount, 2)
            diff_coupon_orders = w2_coupon_orders - w1_coupon_orders
            diff_phone_clicks = w2_phone_clicks - w1_phone_clicks
            diff_avg_price = round(w2_avg_price - w1_avg_price, 2)

            # 差值百分比
            def calc_rate_diff(rate1_str, rate2_str):
                val1 = float(rate1_str.rstrip('%')) if rate1_str != '0%' else 0
                val2 = float(rate2_str.rstrip('%')) if rate2_str != '0%' else 0
                return f"{round(val2 - val1, 1)}%"

            diff_exposure_rate = calc_rate_diff(w1_exposure_rate, w2_exposure_rate)
            diff_order_rate = calc_rate_diff(w1_order_rate, w2_order_rate)

            # 行1: 第一周核销数据
            row1 = [
                shop_name, week1_period,
                round(w1_verify_discount, 2), w1_exposure, w1_visit, w1_exposure_rate,
                w1_order_users, w1_order_coupons, w1_order_rate,
                w1_verify_users, w1_verify_coupons,
                round(w1_order_amount, 2), round(w1_verify_amount, 2),
                w1_coupon_orders, w1_phone_clicks, w1_avg_price
            ]
            ws_summary.append(row1)

            # 行2: 第二周核销数据
            row2 = [
                shop_name, week2_period,
                round(w2_verify_discount, 2), w2_exposure, w2_visit, w2_exposure_rate,
                w2_order_users, w2_order_coupons, w2_order_rate,
                w2_verify_users, w2_verify_coupons,
                round(w2_order_amount, 2), round(w2_verify_amount, 2),
                w2_coupon_orders, w2_phone_clicks, w2_avg_price
            ]
            ws_summary.append(row2)

            # 行3: 差值
            row3 = [
                shop_name, '差值',
                diff_verify_discount, diff_exposure, diff_visit, diff_exposure_rate,
                diff_order_users, diff_order_coupons, diff_order_rate,
                diff_verify_users, diff_verify_coupons,
                diff_order_amount, diff_verify_amount,
                diff_coupon_orders, diff_phone_clicks, diff_avg_price
            ]
            ws_summary.append(row3)

            # 行4: 推广通表头
            header_row = [
                '门店', '数据周期', '推广通花费', '推广通曝光', '推广通点击', '推广通点击均价',
                '推广通订单量', '推广通下单转化率', '推广通查看团购', '推广通查看电话',
                '在线咨询', '地址点击', '门店收藏', '收藏率', '新增好评数', '留评率'
            ]
            ws_summary.append(header_row)

            # 推广通相关数据
            w1_promo_cost = get_val(w1, 'promotion_cost')
            w1_promo_exposure = get_val(w1, 'promotion_exposure')
            w1_promo_clicks = get_val(w1, 'promotion_clicks')
            w1_promo_orders = get_val(w1, 'promotion_orders')
            w1_view_groupbuy = get_val(w1, 'view_groupbuy')
            w1_view_phone = get_val(w1, 'view_phone')
            w1_consult = get_val(w1, 'consult_users')
            w1_address = get_val(w1, 'address_clicks')
            w1_collect = get_val(w1, 'new_collect')
            w1_good_reviews = get_val(w1, 'new_good_reviews')

            w1_click_price = calc_avg_price(w1_promo_cost, w1_promo_clicks)
            w1_promo_rate = f"{calc_rate(w1_promo_orders, w1_promo_clicks)}%"
            w1_collect_rate = f"{calc_rate(w1_collect, w1_visit)}%"
            w1_review_rate = f"{calc_rate(w1_good_reviews, w1_verify_users)}%"

            w2_promo_cost = get_val(w2, 'promotion_cost')
            w2_promo_exposure = get_val(w2, 'promotion_exposure')
            w2_promo_clicks = get_val(w2, 'promotion_clicks')
            w2_promo_orders = get_val(w2, 'promotion_orders')
            w2_view_groupbuy = get_val(w2, 'view_groupbuy')
            w2_view_phone = get_val(w2, 'view_phone')
            w2_consult = get_val(w2, 'consult_users')
            w2_address = get_val(w2, 'address_clicks')
            w2_collect = get_val(w2, 'new_collect')
            w2_good_reviews = get_val(w2, 'new_good_reviews')

            w2_click_price = calc_avg_price(w2_promo_cost, w2_promo_clicks)
            w2_promo_rate = f"{calc_rate(w2_promo_orders, w2_promo_clicks)}%"
            w2_collect_rate = f"{calc_rate(w2_collect, w2_visit)}%"
            w2_review_rate = f"{calc_rate(w2_good_reviews, w2_verify_users)}%"

            # 差值
            diff_promo_cost = round(w2_promo_cost - w1_promo_cost, 2)
            diff_promo_exposure = w2_promo_exposure - w1_promo_exposure
            diff_promo_clicks = w2_promo_clicks - w1_promo_clicks
            diff_click_price = round(w2_click_price - w1_click_price, 2)
            diff_promo_orders = w2_promo_orders - w1_promo_orders
            diff_promo_rate = calc_rate_diff(w1_promo_rate, w2_promo_rate)
            diff_view_groupbuy = w2_view_groupbuy - w1_view_groupbuy
            diff_view_phone = w2_view_phone - w1_view_phone
            diff_consult = w2_consult - w1_consult
            diff_address = w2_address - w1_address
            diff_collect = w2_collect - w1_collect
            diff_collect_rate = calc_rate_diff(w1_collect_rate, w2_collect_rate)
            diff_good_reviews = w2_good_reviews - w1_good_reviews
            diff_review_rate = calc_rate_diff(w1_review_rate, w2_review_rate)

            # 行5: 第一周推广通数据
            row5 = [
                shop_name, week1_period,
                round(w1_promo_cost, 2), w1_promo_exposure, w1_promo_clicks, w1_click_price,
                w1_promo_orders, w1_promo_rate, w1_view_groupbuy, w1_view_phone,
                w1_consult, w1_address, w1_collect, w1_collect_rate,
                w1_good_reviews, w1_review_rate
            ]
            ws_summary.append(row5)

            # 行6: 第二周推广通数据
            row6 = [
                shop_name, week2_period,
                round(w2_promo_cost, 2), w2_promo_exposure, w2_promo_clicks, w2_click_price,
                w2_promo_orders, w2_promo_rate, w2_view_groupbuy, w2_view_phone,
                w2_consult, w2_address, w2_collect, w2_collect_rate,
                w2_good_reviews, w2_review_rate
            ]
            ws_summary.append(row6)

            # 行7: 推广通差值
            row7 = [
                shop_name, '差值',
                diff_promo_cost, diff_promo_exposure, diff_promo_clicks, diff_click_price,
                diff_promo_orders, diff_promo_rate, diff_view_groupbuy, diff_view_phone,
                diff_consult, diff_address, diff_collect, diff_collect_rate,
                diff_good_reviews, diff_review_rate
            ]
            ws_summary.append(row7)

            # 行8: 空行分隔
            ws_summary.append([''] * 16)

            # ==================== 门店详细Sheet（竖向31行）====================
            sheet_name = clean_sheet_name(shop_name)
            if sheet_name in sheet_names_used:
                sheet_names_used[sheet_name] += 1
                sheet_name = f"{sheet_name[:28]}_{sheet_names_used[sheet_name]}"
            else:
                sheet_names_used[sheet_name] = 1

            ws_detail = wb.create_sheet(title=sheet_name)

            # 计算额外指标
            w1_intent_rate = calc_rate(w1_order_users, w1_visit)
            w2_intent_rate = calc_rate(w2_order_users, w2_visit)
            w1_checkin = get_val(w1, 'checkin_count')
            w2_checkin = get_val(w2, 'checkin_count')
            w1_reviews = get_val(w1, 'new_reviews')
            w2_reviews = get_val(w2, 'new_reviews')

            # 构建竖向表格（31行）
            detail_data = [
                [shop_name, '', '', ''],
                ['指标项/时间周期', week1_period, week2_period, '差值（红涨/黑跌）'],
                ['曝光人数：', w1_exposure, w2_exposure, f'=C3-B3'],
                ['访问人数：', w1_visit, w2_visit, f'=C4-B4'],
                ['曝光访问转化率：', f'=B4/B3', f'=C4/C3', f'=C5-B5'],
                ['下单人数：', w1_order_users, w2_order_users, f'=C6-B6'],
                ['核销人数：', w1_verify_users, w2_verify_users, f'=C7-B7'],
                ['意向转化率：', f'=B6/B4', f'=C6/C4', f'=C8-B8'],
                ['下单券数：', w1_order_coupons, w2_order_coupons, f'=C9-B9'],
                ['核销券数：', w1_verify_coupons, w2_verify_coupons, f'=C10-B10'],
                ['下单售价金额：', round(w1_order_amount, 2), round(w2_order_amount, 2), f'=C11-B11'],
                ['核销售价金额：', round(w1_verify_amount, 2), round(w2_verify_amount, 2), f'=C12-B12'],
                ['优惠后核销金额：', round(w1_verify_discount, 2), round(w2_verify_discount, 2), f'=C13-B13'],
                ['客单价：', w1_avg_price, w2_avg_price, f'=C14-B14'],
                ['电话点击：', w1_phone_clicks, w2_phone_clicks, f'=C15-B15'],
                ['地址点击：', w1_address, w2_address, f'=C16-B16'],
                ['在线咨询：', w1_consult, w2_consult, f'=C17-B17'],
                ['门店干预数据', '', '', ''],
                ['新增好评：', w1_good_reviews, w2_good_reviews, f'=C19-B19'],
                ['留评率：', f'=B19/B7', f'=C19/C7', f'=C20-B20'],
                ['门店收藏：', w1_collect, w2_collect, f'=C21-B21'],
                ['收藏率：', f'=B21/B6', f'=C21/C6', f'=C22-B22'],
                ['打卡人数：', w1_checkin, w2_checkin, f'=C23-B23'],
                ['推广通数据', '', '', ''],
                ['推广通订单量', w1_promo_orders, w2_promo_orders, f'=C25-B25'],
                ['推广通花费', round(w1_promo_cost, 2), round(w2_promo_cost, 2), f'=C26-B26'],
                ['推广通曝光（次）', w1_promo_exposure, w2_promo_exposure, f'=C27-B27'],
                ['推广通点击（次）', w1_promo_clicks, w2_promo_clicks, f'=C28-B28'],
                ['推广通点击均价（元）', w1_click_price, w2_click_price, f'=C29-B29'],
                ['查看团购（次）', w1_view_groupbuy, w2_view_groupbuy, f'=C30-B30'],
                ['查看电话（次）', w1_view_phone, w2_view_phone, f'=C31-B31'],
            ]

            for row_data in detail_data:
                ws_detail.append(row_data)

            # 设置详细Sheet样式
            ws_detail.column_dimensions['A'].width = 22
            ws_detail.column_dimensions['B'].width = 18
            ws_detail.column_dimensions['C'].width = 18
            ws_detail.column_dimensions['D'].width = 20

            # 标题样式
            ws_detail['A1'].font = Font(bold=True, size=12)
            ws_detail['A2'].font = Font(bold=True, size=10)

            # 分类标题
            for r in [18, 24]:
                ws_detail.cell(row=r, column=1).font = Font(bold=True, size=10, color="0066CC")

            # 应用边框
            apply_border(ws_detail, 1, len(detail_data), 1, 4)

            # 设置差值列条件格式（红色为上升，黑色为下降）
            red_font = Font(color="FF0000")
            for row_num in range(3, len(detail_data) + 1):
                cell = ws_detail.cell(row=row_num, column=4)
                # 公式单元格会自动计算，这里设置数字格式
                cell.number_format = '0.00;-0.00'

        # 汇总表样式
        for i in range(1, 17):
            ws_summary.column_dimensions[get_column_letter(i)].width = 15
        ws_summary.column_dimensions['A'].width = 40

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row, min_col=1, max_col=16):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # 差值行灰色背景
                if cell.column == 2 and cell.value == '差值':
                    for c in row:
                        c.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

                # 表头行加粗
                if cell.column == 1 and cell.value == '门店':
                    for c in row:
                        c.font = Font(bold=True, size=10)
                        c.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # 保存文件
        if not output_filename:
            output_filename = f"周报 非餐 {week2_start.replace('-', '')}~{week2_end.replace('-', '')} {datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        wb.save(output_filename)
        print(f"✅ 周报生成成功: {output_filename}")
        return output_filename

    finally:
        cursor.close()
        conn.close()


# ==================== 核心功能：生成月报 ====================
def generate_monthly_report(month1_start, month1_end, month2_start, month2_end, output_filename=None):
    """
    生成月报（两个月对比）
    结构与周报完全相同，只是时间跨度从周变为月

    参数:
        month1_start: str, 第一个月开始日期 'YYYY-MM-DD'
        month1_end: str, 第一个月结束日期 'YYYY-MM-DD'
        month2_start: str, 第二个月开始日期 'YYYY-MM-DD'
        month2_end: str, 第二个月结束日期 'YYYY-MM-DD'
        output_filename: str, 输出文件名
    """
    # 生成默认文件名
    if not output_filename:
        output_filename = f"月报 非餐 {month2_start.replace('-', '')}~{month2_end.replace('-', '')} {datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    # 复用周报逻辑
    return generate_weekly_report(month1_start, month1_end, month2_start, month2_end, output_filename)


# ==================== 核心功能：生成自定义报表 ====================
def generate_custom_report(period1_start, period1_end, period2_start, period2_end, shop_ids=None, output_filename=None):
    """
    生成自定义报表（两个自定义时间段对比，支持筛选门店）
    结构与周报相同

    参数:
        period1_start: str, 第一个时期开始日期
        period1_end: str, 第一个时期结束日期
        period2_start: str, 第二个时期开始日期
        period2_end: str, 第二个时期结束日期
        shop_ids: list, 门店ID列表，为空则查询所有门店
        output_filename: str, 输出文件名
    """
    # 获取门店信息映射
    shop_mapping = get_shop_info_mapping()

    conn = CONNECTION_POOL.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 构建 SQL，支持门店筛选
        shop_filter = ""
        if shop_ids:
            shop_filter = f"AND k.shop_id IN ({','.join(map(str, shop_ids))})"

        sql_period = f"""
        SELECT
            k.shop_id,
            k.shop_name,
            k.city,
            SUM(k.verify_after_discount) as verify_after_discount,
            SUM(k.exposure_users) as exposure_users,
            SUM(k.visit_users) as visit_users,
            SUM(k.order_users) as order_users,
            SUM(k.order_coupon_count) as order_coupon_count,
            SUM(k.verify_person_count) as verify_users,
            SUM(k.verify_coupon_count) as verify_coupon_count,
            SUM(k.order_sale_amount) as order_sale_amount,
            SUM(k.verify_sale_amount) as verify_sale_amount,
            SUM(k.coupon_pay_order_count) as coupon_orders,
            SUM(p.view_phone_count) as phone_clicks,
            SUM(k.promotion_cost) as promotion_cost,
            SUM(k.promotion_exposure_count) as promotion_exposure,
            SUM(k.promotion_click_count) as promotion_clicks,
            SUM(p.order_count) as promotion_orders,
            SUM(p.view_groupbuy_count) as view_groupbuy,
            SUM(p.view_phone_count) as view_phone,
            SUM(k.consult_users) as consult_users,
            SUM(p.view_address_count) as address_clicks,
            SUM(k.new_collect_users) as new_collect,
            SUM(k.new_good_review_count) as new_good_reviews,
            SUM(k.new_review_count) as new_reviews,
            SUM(s.checkin_count) as checkin_count
        FROM kewen_daily_report k
        LEFT JOIN promotion_daily_report p
            ON k.shop_id = p.shop_id AND k.report_date = p.report_date
        LEFT JOIN store_stats s
            ON k.shop_id = s.store_id AND k.report_date = s.date
        WHERE k.report_date BETWEEN %s AND %s {shop_filter}
        GROUP BY k.shop_id, k.shop_name, k.city
        ORDER BY k.shop_id
        """

        # 获取两个时期的数据
        cursor.execute(sql_period, (period1_start, period1_end))
        period1_data = {row['shop_id']: row for row in cursor.fetchall()}

        cursor.execute(sql_period, (period2_start, period2_end))
        period2_data = {row['shop_id']: row for row in cursor.fetchall()}

        all_shop_ids_set = set(period1_data.keys()) | set(period2_data.keys())

        if not all_shop_ids_set:
            print("警告：没有找到数据")
            return None

        # 创建 Excel
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "自定义报表"

        # 格式化日期周期
        period1_str = f"{datetime.strptime(period1_start, '%Y-%m-%d').strftime('%Y.%m.%d')}-{datetime.strptime(period1_end, '%Y-%m-%d').strftime('%Y.%m.%d')}"
        period2_str = f"{datetime.strptime(period2_start, '%Y-%m-%d').strftime('%Y.%m.%d')}-{datetime.strptime(period2_end, '%Y-%m-%d').strftime('%Y.%m.%d')}"

        # 辅助函数
        def get_val(data, key, default=0):
            return data.get(key, default) if data else default

        def calc_rate(numerator, denominator):
            if denominator and denominator > 0:
                return round(numerator / denominator * 100, 1)
            return 0

        def calc_avg_price(total, count):
            if count and count > 0:
                return round(total / count, 2)
            return 0

        # 用于处理重名 Sheet
        sheet_names_used = {}
        seq_num = 1

        for shop_id in sorted(all_shop_ids_set):
            p1 = period1_data.get(shop_id, {})
            p2 = period2_data.get(shop_id, {})
            shop_name = p2.get('shop_name') or p1.get('shop_name', '未知门店')

            # 从映射中获取运营、城市、销售
            shop_id_str = str(shop_id)
            shop_info = shop_mapping.get(shop_id_str, {})
            operator = shop_info.get('operator', '')
            sales = shop_info.get('sales', '')
            city = shop_info.get('city', '')

            # ==================== 汇总表数据（与周报相同的8行结构）====================
            # 时期1指标
            p1_verify_discount = get_val(p1, 'verify_after_discount')
            p1_exposure = get_val(p1, 'exposure_users')
            p1_visit = get_val(p1, 'visit_users')
            p1_order_users = get_val(p1, 'order_users')
            p1_order_coupons = get_val(p1, 'order_coupon_count')
            p1_verify_users = get_val(p1, 'verify_users')
            p1_verify_coupons = get_val(p1, 'verify_coupon_count')
            p1_order_amount = get_val(p1, 'order_sale_amount')
            p1_verify_amount = get_val(p1, 'verify_sale_amount')
            p1_coupon_orders = get_val(p1, 'coupon_orders')
            p1_phone_clicks = get_val(p1, 'phone_clicks')

            p1_exposure_rate = f"{calc_rate(p1_visit, p1_exposure)}%"
            p1_order_rate = f"{calc_rate(p1_order_users, p1_visit)}%"
            p1_avg_price = calc_avg_price(p1_verify_discount, p1_verify_users)

            # 时期2指标
            p2_verify_discount = get_val(p2, 'verify_after_discount')
            p2_exposure = get_val(p2, 'exposure_users')
            p2_visit = get_val(p2, 'visit_users')
            p2_order_users = get_val(p2, 'order_users')
            p2_order_coupons = get_val(p2, 'order_coupon_count')
            p2_verify_users = get_val(p2, 'verify_users')
            p2_verify_coupons = get_val(p2, 'verify_coupon_count')
            p2_order_amount = get_val(p2, 'order_sale_amount')
            p2_verify_amount = get_val(p2, 'verify_sale_amount')
            p2_coupon_orders = get_val(p2, 'coupon_orders')
            p2_phone_clicks = get_val(p2, 'phone_clicks')

            p2_exposure_rate = f"{calc_rate(p2_visit, p2_exposure)}%"
            p2_order_rate = f"{calc_rate(p2_order_users, p2_visit)}%"
            p2_avg_price = calc_avg_price(p2_verify_discount, p2_verify_users)

            # 差值计算
            diff_verify_discount = round(p2_verify_discount - p1_verify_discount, 2)
            diff_exposure = p2_exposure - p1_exposure
            diff_visit = p2_visit - p1_visit

            def calc_rate_diff(rate1_str, rate2_str):
                val1 = float(rate1_str.rstrip('%')) if rate1_str != '0%' else 0
                val2 = float(rate2_str.rstrip('%')) if rate2_str != '0%' else 0
                return f"{round(val2 - val1, 1)}%"

            diff_exposure_rate = calc_rate_diff(p1_exposure_rate, p2_exposure_rate)
            diff_order_users = p2_order_users - p1_order_users
            diff_order_coupons = p2_order_coupons - p1_order_coupons
            diff_order_rate = calc_rate_diff(p1_order_rate, p2_order_rate)
            diff_verify_users = p2_verify_users - p1_verify_users
            diff_verify_coupons = p2_verify_coupons - p1_verify_coupons
            diff_order_amount = round(p2_order_amount - p1_order_amount, 2)
            diff_verify_amount = round(p2_verify_amount - p1_verify_amount, 2)
            diff_coupon_orders = p2_coupon_orders - p1_coupon_orders
            diff_phone_clicks = p2_phone_clicks - p1_phone_clicks
            diff_avg_price = round(p2_avg_price - p1_avg_price, 2)

            # 推广通数据
            p1_promo_cost = get_val(p1, 'promotion_cost')
            p1_promo_exposure = get_val(p1, 'promotion_exposure')
            p1_promo_clicks = get_val(p1, 'promotion_clicks')
            p1_promo_orders = get_val(p1, 'promotion_orders')
            p1_view_groupbuy = get_val(p1, 'view_groupbuy')
            p1_view_phone = get_val(p1, 'view_phone')
            p1_consult = get_val(p1, 'consult_users')
            p1_address = get_val(p1, 'address_clicks')
            p1_collect = get_val(p1, 'new_collect')
            p1_good_reviews = get_val(p1, 'new_good_reviews')
            p1_click_price = calc_avg_price(p1_promo_cost, p1_promo_clicks)
            p1_promo_rate = f"{calc_rate(p1_promo_orders, p1_promo_clicks)}%"
            p1_collect_rate = f"{calc_rate(p1_collect, p1_visit)}%"
            p1_review_rate = f"{calc_rate(p1_good_reviews, p1_verify_users)}%"

            p2_promo_cost = get_val(p2, 'promotion_cost')
            p2_promo_exposure = get_val(p2, 'promotion_exposure')
            p2_promo_clicks = get_val(p2, 'promotion_clicks')
            p2_promo_orders = get_val(p2, 'promotion_orders')
            p2_view_groupbuy = get_val(p2, 'view_groupbuy')
            p2_view_phone = get_val(p2, 'view_phone')
            p2_consult = get_val(p2, 'consult_users')
            p2_address = get_val(p2, 'address_clicks')
            p2_collect = get_val(p2, 'new_collect')
            p2_good_reviews = get_val(p2, 'new_good_reviews')
            p2_click_price = calc_avg_price(p2_promo_cost, p2_promo_clicks)
            p2_promo_rate = f"{calc_rate(p2_promo_orders, p2_promo_clicks)}%"
            p2_collect_rate = f"{calc_rate(p2_collect, p2_visit)}%"
            p2_review_rate = f"{calc_rate(p2_good_reviews, p2_verify_users)}%"

            diff_promo_cost = round(p2_promo_cost - p1_promo_cost, 2)
            diff_promo_exposure = p2_promo_exposure - p1_promo_exposure
            diff_promo_clicks = p2_promo_clicks - p1_promo_clicks
            diff_click_price = round(p2_click_price - p1_click_price, 2)
            diff_promo_orders = p2_promo_orders - p1_promo_orders
            diff_promo_rate = calc_rate_diff(p1_promo_rate, p2_promo_rate)
            diff_view_groupbuy = p2_view_groupbuy - p1_view_groupbuy
            diff_view_phone = p2_view_phone - p1_view_phone
            diff_consult = p2_consult - p1_consult
            diff_address = p2_address - p1_address
            diff_collect = p2_collect - p1_collect
            diff_collect_rate = calc_rate_diff(p1_collect_rate, p2_collect_rate)
            diff_good_reviews = p2_good_reviews - p1_good_reviews
            diff_review_rate = calc_rate_diff(p1_review_rate, p2_review_rate)

            # 第一行：时期1核销数据
            row1 = [
                seq_num, operator, city, sales, shop_name, period1_str,
                round(p1_verify_discount, 2), p1_exposure, p1_visit, p1_exposure_rate,
                p1_order_users, p1_order_coupons, p1_order_rate,
                p1_verify_users, p1_verify_coupons,
                round(p1_order_amount, 2), round(p1_verify_amount, 2),
                p1_coupon_orders, p1_phone_clicks, p1_avg_price,
                round(p1_promo_cost, 2), p1_promo_exposure, p1_promo_clicks, p1_click_price,
                p1_promo_orders, p1_promo_rate, p1_view_groupbuy, p1_view_phone,
                p1_consult, p1_address, p1_collect, p1_collect_rate,
                p1_good_reviews, p1_review_rate
            ]
            ws_summary.append(row1)

            # 第二行：时期2数据
            row2 = [
                seq_num, operator, city, sales, shop_name, period2_str,
                round(p2_verify_discount, 2), p2_exposure, p2_visit, p2_exposure_rate,
                p2_order_users, p2_order_coupons, p2_order_rate,
                p2_verify_users, p2_verify_coupons,
                round(p2_order_amount, 2), round(p2_verify_amount, 2),
                p2_coupon_orders, p2_phone_clicks, p2_avg_price,
                round(p2_promo_cost, 2), p2_promo_exposure, p2_promo_clicks, p2_click_price,
                p2_promo_orders, p2_promo_rate, p2_view_groupbuy, p2_view_phone,
                p2_consult, p2_address, p2_collect, p2_collect_rate,
                p2_good_reviews, p2_review_rate
            ]
            ws_summary.append(row2)

            # 第三行：差值
            row3 = [
                seq_num, operator, city, sales, shop_name, '差值',
                diff_verify_discount, diff_exposure, diff_visit, diff_exposure_rate,
                diff_order_users, diff_order_coupons, diff_order_rate,
                diff_verify_users, diff_verify_coupons,
                diff_order_amount, diff_verify_amount,
                diff_coupon_orders, diff_phone_clicks, diff_avg_price,
                diff_promo_cost, diff_promo_exposure, diff_promo_clicks, diff_click_price,
                diff_promo_orders, diff_promo_rate, diff_view_groupbuy, diff_view_phone,
                diff_consult, diff_address, diff_collect, diff_collect_rate,
                diff_good_reviews, diff_review_rate
            ]
            ws_summary.append(row3)

            # 第四行：表头（重复）
            header = [
                '序号', '运营', '城市', '销售', '门店', '数据周期',
                '优惠后核销额', '曝光人数', '访问人数', '曝光访问转化率',
                '下单人数', '下单券数', '下单转化率', '核销人数', '核销券数',
                '下单售价金额', '核销售价金额', '优惠码订单', '电话点击', '客单价',
                '推广通花费', '推广通曝光', '推广通点击', '推广通点击均价',
                '推广通订单量', '推广通下单转化率', '推广通查看团购', '推广通查看电话',
                '在线咨询', '地址点击', '门店收藏', '收藏率', '新增好评数', '留评率'
            ]
            ws_summary.append(header)

            # ==================== 门店详细Sheet（竖向31行）====================
            sheet_name = clean_sheet_name(shop_name)
            if sheet_name in sheet_names_used:
                sheet_names_used[sheet_name] += 1
                sheet_name = f"{sheet_name[:28]}_{sheet_names_used[sheet_name]}"
            else:
                sheet_names_used[sheet_name] = 1

            ws_detail = wb.create_sheet(title=sheet_name)

            # 额外指标
            p1_checkin = get_val(p1, 'checkin_count')
            p2_checkin = get_val(p2, 'checkin_count')

            # 构建竖向表格（31行）
            detail_data = [
                [shop_name, '', '', ''],
                ['指标项/时间周期', period1_str, period2_str, '差值（红涨/黑跌）'],
                ['曝光人数：', p1_exposure, p2_exposure, f'=C3-B3'],
                ['访问人数：', p1_visit, p2_visit, f'=C4-B4'],
                ['曝光访问转化率：', f'=B4/B3', f'=C4/C3', f'=C5-B5'],
                ['下单人数：', p1_order_users, p2_order_users, f'=C6-B6'],
                ['核销人数：', p1_verify_users, p2_verify_users, f'=C7-B7'],
                ['意向转化率：', f'=B6/B4', f'=C6/C4', f'=C8-B8'],
                ['下单券数：', p1_order_coupons, p2_order_coupons, f'=C9-B9'],
                ['核销券数：', p1_verify_coupons, p2_verify_coupons, f'=C10-B10'],
                ['下单售价金额：', round(p1_order_amount, 2), round(p2_order_amount, 2), f'=C11-B11'],
                ['核销售价金额：', round(p1_verify_amount, 2), round(p2_verify_amount, 2), f'=C12-B12'],
                ['优惠后核销金额：', round(p1_verify_discount, 2), round(p2_verify_discount, 2), f'=C13-B13'],
                ['客单价：', p1_avg_price, p2_avg_price, f'=C14-B14'],
                ['电话点击：', p1_phone_clicks, p2_phone_clicks, f'=C15-B15'],
                ['地址点击：', p1_address, p2_address, f'=C16-B16'],
                ['在线咨询：', p1_consult, p2_consult, f'=C17-B17'],
                ['门店干预数据', '', '', ''],
                ['新增好评：', p1_good_reviews, p2_good_reviews, f'=C19-B19'],
                ['留评率：', f'=B19/B7', f'=C19/C7', f'=C20-B20'],
                ['门店收藏：', p1_collect, p2_collect, f'=C21-B21'],
                ['收藏率：', f'=B21/B6', f'=C21/C6', f'=C22-B22'],
                ['打卡人数：', p1_checkin, p2_checkin, f'=C23-B23'],
                ['推广通数据', '', '', ''],
                ['推广通订单量', p1_promo_orders, p2_promo_orders, f'=C25-B25'],
                ['推广通花费', round(p1_promo_cost, 2), round(p2_promo_cost, 2), f'=C26-B26'],
                ['推广通曝光（次）', p1_promo_exposure, p2_promo_exposure, f'=C27-B27'],
                ['推广通点击（次）', p1_promo_clicks, p2_promo_clicks, f'=C28-B28'],
                ['推广通点击均价（元）', p1_click_price, p2_click_price, f'=C29-B29'],
                ['查看团购（次）', p1_view_groupbuy, p2_view_groupbuy, f'=C30-B30'],
                ['查看电话（次）', p1_view_phone, p2_view_phone, f'=C31-B31'],
            ]

            for row_data in detail_data:
                ws_detail.append(row_data)

            # 设置详细Sheet样式
            ws_detail.column_dimensions['A'].width = 22
            ws_detail.column_dimensions['B'].width = 18
            ws_detail.column_dimensions['C'].width = 18
            ws_detail.column_dimensions['D'].width = 20

            # 标题样式
            ws_detail['A1'].font = Font(bold=True, size=12)
            ws_detail['A2'].font = Font(bold=True, size=10)

            # 分类标题
            for r in [18, 24]:
                ws_detail.cell(row=r, column=1).font = Font(bold=True, size=10, color="0066CC")

            # 应用边框
            apply_border(ws_detail, 1, len(detail_data), 1, 4)

            seq_num += 1

        # 汇总表样式
        for i in range(1, 35):
            ws_summary.column_dimensions[get_column_letter(i)].width = 12
        ws_summary.column_dimensions['E'].width = 40

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row, min_col=1, max_col=34):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # 差值行灰色背景
                if cell.column == 6 and cell.value == '差值':
                    for c in row:
                        c.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

                # 表头行加粗
                if cell.value == '序号' and ws_summary.cell(cell.row, 2).value == '运营':
                    for c in row:
                        c.font = Font(bold=True, size=10)
                        c.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # 保存文件
        if not output_filename:
            shop_count = len(all_shop_ids_set)
            output_filename = f"自定义 {shop_count}家门店非餐 {period2_start.replace('-', '')}~{period2_end.replace('-', '')} {datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        wb.save(output_filename)
        print(f"✅ 自定义报表生成成功: {output_filename}")
        return output_filename

    finally:
        cursor.close()
        conn.close()


# ==================== 主程序示例 ====================
if __name__ == "__main__":
    print("=" * 60)
    print("江鑫数据报表生成系统")
    print("=" * 60)

    # 示例：生成日报
    print("\n【示例1】生成日报")
    try:
        generate_daily_report('2025-12-12')
    except Exception as e:
        print(f"❌ 日报生成失败: {e}")

    # 示例：生成周报
    print("\n【示例2】生成周报")
    try:
        generate_weekly_report(
            week1_start='2025-11-10',
            week1_end='2025-11-16',
            week2_start='2025-11-17',
            week2_end='2025-11-23'
        )
    except Exception as e:
        print(f"❌ 周报生成失败: {e}")

    # 示例：生成月报
    print("\n【示例3】生成月报")
    try:
        generate_monthly_report(
            month1_start='2025-09-01',
            month1_end='2025-09-30',
            month2_start='2025-10-01',
            month2_end='2025-10-31'
        )
    except Exception as e:
        print(f"❌ 月报生成失败: {e}")

    # 示例：生成自定义报表
    print("\n【示例4】生成自定义报表")
    try:
        generate_custom_report(
            period1_start='2025-10-25',
            period1_end='2025-11-09',
            period2_start='2025-11-10',
            period2_end='2025-11-25',
            shop_ids=None  # None表示所有门店，也可以传入 [shop_id1, shop_id2, ...]
        )
    except Exception as e:
        print(f"❌ 自定义报表生成失败: {e}")

    print("\n" + "=" * 60)
    print("所有报表生成完成！")
    print("=" * 60)
